#!/usr/bin/env python3
"""
Menghasilkan src/mocks/data/{program,kegiatan,subkegiatan,aktifitas}.ts
dan src/mocks/data/penugasan-awal.ts dari sumber Excel asli.

Kenapa: roadmap-refactor-ui.md Fase 0 minta data dummy mengikuti bentuk
ERD TUJUAN, bukan dikarang bebas. File ini membaca struktur baris/kolom
asli "Operasional Indikator Kinerja BAPPERIDA.xlsx" dan memetakannya ke
MASTER_PROGRAM / MASTER_KEGIATAN / MASTER_INDIKATOR_UTAMA / MASTER_INDIKATOR
(lihat docs/erd.md), supaya 78 subkegiatan dan aktifitasnya di layar demo
adalah nama asli BAPPERIDA, bukan "Subkegiatan 1..78".

Jalankan ulang tiap kali file Excel sumber berubah:
    python3 scripts/seed_from_xlsx.py

Yang TIDAK diambil dari Excel (tidak ada di sana / nilainya basi):
- BOBOT_TARGET aktifitas pendukung: Excel selalu menulis 0.1 per baris
  apa pun jumlah pendukungnya (kadang totalnya jadi 40%, bukan 30%).
  Bobot yang benar dihitung ulang di runtime oleh src/lib/bobot.ts
  (30% ÷ jumlah aktifitas pendukung AKTIF), sesuai aturan REFACTOR.md
  "bobot tidak pernah jadi input/nilai tersimpan yang dipercaya mentah".
- Realisasi: template Excel ini kosong (semua Bobot Realisasi = 0),
  jadi bukan sumber untuk data realisasi contoh. Itu digenerate terpisah
  oleh src/mocks/db.ts memakai PRNG berbenih (deterministik).
"""
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Perlu openpyxl: pip install openpyxl (atau jalankan lewat python3 "
        "yang sudah punya paket ini)."
    )

REPO_ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = (
    REPO_ROOT.parent / "docs" / "Operasional Indikator Kinerja BAPPERIDA.xlsx"
)
OUT_DIR = REPO_ROOT / "src" / "mocks" / "data"
SHEET_NAME = "OPERA INK"

BIDANG_MAP = {
    "P2EPD": "P2EPD",
    "SEKRETARIAT": "SEKRETARIAT",
    "SEKERTARIAT": "SEKRETARIAT",  # salah ketik di sumber, disamakan
    "PIK": "PIK",
    "RINOVA": "RINOVA",
    "PERENCANAAN": "SUBBAG_PERENCANAAN",
    "SUBBAG PERENCANAAN": "SUBBAG_PERENCANAAN",
    "SUBBAGIAN PERENCANAAN": "SUBBAG_PERENCANAAN",
    "PPM": "PPM",
    "KEUANGAN": "KEUANGAN",
}

GENERATED_HEADER = """// ⚠️ BERKAS HASIL GENERATE — jangan edit manual.
// Sumber: "Operasional Indikator Kinerja BAPPERIDA.xlsx" (sheet "{sheet}").
// Regenerasi: python3 scripts/seed_from_xlsx.py
"""


def normalize_bidang(raw):
    if not isinstance(raw, str):
        return None
    s = raw.strip().upper()
    s = re.sub(r"^BID\.?\s+", "", s)
    s = re.sub(r"^BIDANG\s+", "", s)
    return BIDANG_MAP.get(s.strip())


def fmt_segment(v):
    """5.0 -> '5', 2.01 -> '2.01' — meniru format kode Dinas di Excel."""
    if v is None or v == "":
        return None
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return ("%g" % v).replace("e-0", "e-")
    return str(v)


def build_kode(*segments):
    parts = [fmt_segment(s) for s in segments]
    return ".".join(p for p in parts if p)


def split_target(raw):
    """'2 Dokumen' -> (2.0, 'Dokumen'); 6 -> (6.0, None)."""
    if raw is None:
        return 0.0, None
    if isinstance(raw, (int, float)):
        return float(raw), None
    m = re.match(r"\s*([\d.,]+)\s*(.*)", str(raw))
    if not m:
        return 0.0, None
    num = float(m.group(1).replace(",", "."))
    satuan = m.group(2).strip() or None
    return num, satuan


def esc(s):
    if s is None:
        return "null"
    return json.dumps(str(s).strip(), ensure_ascii=False)


def parse_workbook():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    def cell(r, c):
        return ws.cell(r, c).value

    def depth_of(r):
        kode = [cell(r, c) for c in range(1, 6)]
        return sum(1 for k in kode if k not in (None, ""))

    urusans, programs, kegiatans, subs = [], [], [], []
    urusan_seq = prog_seq = keg_seq = sub_seq = 0
    r, max_r = 5, ws.max_row

    while r <= max_r:
        d = depth_of(r)

        # Kedalaman 2 = URUSAN (mis. "5.0.1.0 PERENCANAAN"). Excel menandainya
        # dengan label "KINERJA URUSAN" di kolom 12. Dipakai sebagai induk
        # pengelompokan program di layar Capaian Program.
        if d == 2:
            urusan_seq += 1
            urusans.append(
                {
                    "id": urusan_seq,
                    "kode": build_kode(cell(r, 1), cell(r, 2)),
                    "nama": (cell(r, 6) or "").strip(),
                }
            )
            r += 1
            continue

        if d == 3:
            prog_seq += 1
            programs.append(
                {
                    "id": prog_seq,
                    "urusanId": urusan_seq,
                    "kode": build_kode(cell(r, 1), cell(r, 2), cell(r, 3)),
                    "nama": (cell(r, 6) or "").strip(),
                }
            )
            r += 1
            continue

        if d == 4:
            keg_seq += 1
            kegiatans.append(
                {
                    "id": keg_seq,
                    "programId": prog_seq,
                    "kode": build_kode(cell(r, 1), cell(r, 2), cell(r, 3), cell(r, 4)),
                    "nama": (cell(r, 6) or "").strip(),
                }
            )
            r += 1
            continue

        if d == 5:
            sub_seq += 1
            nama_sub = (cell(r, 6) or "").strip()
            kode_sub = build_kode(
                cell(r, 1), cell(r, 2), cell(r, 3), cell(r, 4), cell(r, 5)
            )
            indikator_kinerja = (cell(r, 11) or "").strip()
            target_kinerja, satuan_kinerja = split_target(cell(r, 12))

            sub = {
                "id": sub_seq,
                "kegiatanId": keg_seq,
                "kode": kode_sub,
                "nama": nama_sub,
                "indikatorKinerja": indikator_kinerja,
                "targetKinerja": target_kinerja,
                "satuanKinerja": satuan_kinerja or "Kegiatan",
                "bidang": None,
                "aktifitasUtama": None,
                "aktifitasPendukung": [],
            }

            rr = r + 1
            if (
                rr <= max_r
                and isinstance(cell(rr, 6), str)
                and cell(rr, 6).strip().startswith("Aktifitas Utama")
            ):
                sub["bidang"] = normalize_bidang(cell(rr, 11))
                rr += 1
                while rr <= max_r and not (
                    isinstance(cell(rr, 6), str)
                    and cell(rr, 6).strip().startswith("Aktifitas Pendukung")
                ):
                    if cell(rr, 6) and cell(rr, 9) is not None:
                        nama_u = (cell(rr, 6) or "").strip()
                        target_u, satuan_u = split_target(cell(rr, 7))
                        sub["aktifitasUtama"] = {
                            "nama": nama_u,
                            "target": target_u,
                            "satuan": satuan_u or sub["satuanKinerja"],
                        }
                    rr += 1

                if rr <= max_r:
                    rr += 1  # lewati header "Aktifitas Pendukung..."
                    while rr <= max_r and cell(rr, 5) == "*":
                        nama_p = (cell(rr, 6) or "").strip()
                        if nama_p and nama_p != "Tambahkan Indikator Pendukung Lainnya":
                            target_p, satuan_p = split_target(cell(rr, 7))
                            sub["aktifitasPendukung"].append(
                                {
                                    "nama": nama_p,
                                    "target": target_p,
                                    "satuan": satuan_p or "Kegiatan",
                                }
                            )
                        rr += 1

            subs.append(sub)
            r = rr if rr > r else r + 1
            continue

        r += 1

    return urusans, programs, kegiatans, subs


def write_urusan(urusans):
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append(
        "// MASTER_URUSAN — tingkat di atas program (Excel: label "
        '"KINERJA URUSAN").\n'
        "// Belum ada di docs/erd.md; dipakai sebagai pengelompokan di layar\n"
        "// Capaian Program. Capaian per urusan BELUM punya rumus — lihat\n"
        "// docs/keputusan-terbuka.md butir 2.\n"
    )
    lines.append('import type { Urusan } from "@/services/types"\n')
    lines.append("export const urusans: Urusan[] = [")
    for u in urusans:
        lines.append(
            f'  {{ id: {u["id"]}, kodeUrusan: {esc(u["kode"])}, '
            f'namaUrusan: {esc(u["nama"])} }},'
        )
    lines.append("]\n")
    (OUT_DIR / "urusan.ts").write_text("\n".join(lines), encoding="utf-8")


def write_program(programs):
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append('import type { Program } from "@/services/types"')
    lines.append('import { DOKUMEN_PROGRAM_AWAL } from "./dokumen"\n')
    lines.append("export const programs: Program[] = [")
    for p in programs:
        lines.append(
            f'  {{ id: {p["id"]}, dokumenId: DOKUMEN_PROGRAM_AWAL, '
            f'urusanId: {p["urusanId"]}, '
            f'kodeProgram: {esc(p["kode"])}, '
            f'namaProgram: {esc(p["nama"])}, flagActive: true }},'
        )
    lines.append("]\n")
    (OUT_DIR / "program.ts").write_text("\n".join(lines), encoding="utf-8")


def write_kegiatan(kegiatans):
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append('import type { Kegiatan } from "@/services/types"\n')
    lines.append("export const kegiatans: Kegiatan[] = [")
    for k in kegiatans:
        lines.append(
            f'  {{ id: {k["id"]}, programId: {k["programId"]}, '
            f'kodeKegiatan: {esc(k["kode"])}, namaKegiatan: {esc(k["nama"])}, '
            f"flagActive: true }},"
        )
    lines.append("]\n")
    (OUT_DIR / "kegiatan.ts").write_text("\n".join(lines), encoding="utf-8")


def write_subkegiatan(subs):
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append('import type { IndikatorUtama } from "@/services/types"\n')
    lines.append("export const subkegiatans: IndikatorUtama[] = [")
    for s in subs:
        lines.append(
            f'  {{\n'
            f'    id: {s["id"]},\n'
            f'    kegiatanId: {s["kegiatanId"]},\n'
            f'    kodeIndikatorUtama: {esc(s["kode"])},\n'
            f'    namaIndikatorUtama: {esc(s["nama"])},\n'
            f'    indikatorKinerja: {esc(s["indikatorKinerja"])},\n'
            f'    targetKinerja: {s["targetKinerja"]},\n'
            f'    satuanKinerja: {esc(s["satuanKinerja"])},\n'
            f'    outputKinerja: null,\n'
            f'    flagActive: true,\n'
            f'  }},'
        )
    lines.append("]\n")
    (OUT_DIR / "subkegiatan.ts").write_text("\n".join(lines), encoding="utf-8")


def write_aktifitas(subs):
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append(
        '// Katalog aktifitas (utama + pendukung) per subkegiatan.\n'
        "// BOBOT_TARGET tidak disimpan di sini dengan sengaja — dihitung\n"
        "// runtime oleh src/lib/bobot.ts, lihat catatan di kepala skrip ini.\n"
    )
    lines.append('import type { Indikator } from "@/services/types"\n')
    lines.append("export const aktifitas: Indikator[] = [")
    aid = 0
    for s in subs:
        urutan = 0
        if s["aktifitasUtama"]:
            aid += 1
            urutan += 1
            a = s["aktifitasUtama"]
            lines.append(
                f'  {{ id: {aid}, indikatorUtamaId: {s["id"]}, '
                f'kodeIndikator: {esc(s["kode"] + ".U")}, '
                f'namaIndikator: {esc(a["nama"])}, tipeAktifitas: "UTAMA", '
                f'satuan: {esc(a["satuan"])}, targetAnjuran: {a["target"]}, '
                f'urutan: {urutan}, flagActive: true }},'
            )
        for i, a in enumerate(s["aktifitasPendukung"], start=1):
            aid += 1
            urutan += 1
            lines.append(
                f'  {{ id: {aid}, indikatorUtamaId: {s["id"]}, '
                f'kodeIndikator: {esc(s["kode"] + f".P{i}")}, '
                f'namaIndikator: {esc(a["nama"])}, tipeAktifitas: "PENDUKUNG", '
                f'satuan: {esc(a["satuan"])}, targetAnjuran: {a["target"]}, '
                f'urutan: {urutan}, flagActive: true }},'
            )
    lines.append("]\n")
    (OUT_DIR / "aktifitas.ts").write_text("\n".join(lines), encoding="utf-8")


def write_penugasan_awal(subs):
    """Pemetaan indikatorUtamaId -> kode bidang, dari kolom BIDANG di Excel.
    Dipakai HANYA sebagai benih TRANS_SUBKEGIATAN_BIDANG periode aktif di
    src/mocks/db.ts — bukan bagian dari master, sesuai prinsip ERD bahwa
    penugasan bidang adalah data transaksi per periode, bukan katalog."""
    lines = [GENERATED_HEADER.format(sheet=SHEET_NAME)]
    lines.append(
        "// indikatorUtamaId -> kode MASTER_BIDANG, dipakai db.ts untuk\n"
        "// membangun TRANS_SUBKEGIATAN_BIDANG periode yang sedang OPEN.\n"
    )
    lines.append(
        "export const penugasanAwal: Record<number, string> = {"
    )
    for s in subs:
        lines.append(f'  {s["id"]}: {esc(s["bidang"])},')
    lines.append("}\n")
    (OUT_DIR / "penugasan-awal.ts").write_text("\n".join(lines), encoding="utf-8")


def main():
    if not XLSX_PATH.exists():
        sys.exit(f"Excel sumber tidak ditemukan: {XLSX_PATH}")
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    urusans, programs, kegiatans, subs = parse_workbook()

    unresolved = [s["kode"] for s in subs if s["bidang"] is None]
    if unresolved:
        sys.exit(
            "Baris subkegiatan tanpa bidang yang bisa dipetakan: "
            + ", ".join(unresolved)
            + " — perbarui BIDANG_MAP di skrip ini."
        )

    tanpa_urusan = [p["kode"] for p in programs if not p["urusanId"]]
    if tanpa_urusan:
        sys.exit(
            "Program tanpa urusan induk: "
            + ", ".join(tanpa_urusan)
            + " — baris urusan (kedalaman kode 2) hilang di Excel?"
        )

    write_urusan(urusans)
    write_program(programs)
    write_kegiatan(kegiatans)
    write_subkegiatan(subs)
    write_aktifitas(subs)
    write_penugasan_awal(subs)

    n_aktifitas = sum(1 + len(s["aktifitasPendukung"]) for s in subs)
    print(
        f"OK — {len(urusans)} urusan, {len(programs)} program, "
        f"{len(kegiatans)} kegiatan, "
        f"{len(subs)} subkegiatan, {n_aktifitas} aktifitas ditulis ke "
        f"{OUT_DIR.relative_to(REPO_ROOT)}/"
    )


if __name__ == "__main__":
    main()
